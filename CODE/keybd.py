# -*- coding: utf-8 -*-
#####################################################
#    Project:     RollColl with RFID on RPi         #
#    Programmer:  ---- -----                        #
#    Date:        2017 Jun 09                       #
#    For:         ------- University                #
#####################################################

# import needed Packages
from tkinter import font
from tkinter import *
import jdatetime
import serial
from os.path import exists
from openpyxl import Workbook
from openpyxl import load_workbook


# Public Variables
main_page_bool = True
add_page_bool = False
today = jdatetime.date.today()
person = []
row = 0 + 2
file_list = open("variables.txt")
file = file_list.read()
row = file[4:]
row = int(row)
file_list.close()
# Configure Serial Port
rf = serial.Serial("COM3", 9600, timeout=0.1)

# Create and Configure GUI with Tkinter
root = Tk()

# -- Get Screen Resolution and Create Full screen GUI
screen_x = root.winfo_screenwidth()
screen_y = root.winfo_screenheight()
root.overrideredirect(True)
root.geometry("{0}x{1}+0+0".format(root.winfo_screenwidth(), root.winfo_screenheight()))

# -- Create Canvas for Design GUI
w = Canvas(root, width=screen_x, height=screen_y)
w.grid(row=0, column=0)

# -- Create Fonts and use in GUI
# ---- Persian Fonts
list_font = font.Font(family="B Titr", size=20, weight="bold")
time_font = font.Font(family="B Titr", size=130, weight="bold")
date_font = font.Font(family="B Titr", size=40, weight="bold")
name_font = font.Font(family="Times New Roman", size=24, weight="bold")
info_font = font.Font(family="Times New Roman", size=30, weight="bold")
# -- Put a High Resolution in GUI as Background
background_image = PhotoImage(file="images\\background_1920.png")
w.create_image(0, 0, anchor=NW, image=background_image)
Logo_image = PhotoImage(file="images\\Logo.png")
Logo = w.create_image(screen_x/2, screen_y/6, anchor=CENTER, image=Logo_image)
# -- Set Time Variable
TIME_img = w.create_text(screen_x/2, screen_y/2,
                         text=jdatetime.datetime.now().strftime("%H:%M:%S"),
                         font=time_font, fill='white')
DAY_img = w.create_text(12*screen_x/20, 7*screen_y/10,
                         text=jdatetime.datetime.now().strftime("%d"),
                         font=date_font, fill='white')
YEAR_img = w.create_text(8*screen_x/20, 7 * screen_y / 10,
                         text=jdatetime.datetime.now().strftime("%Y"),
                         font=date_font, fill='white')
months_file = "images\Months\\" + str(jdatetime.date.today())[5:7] + ".png"
month_img = PhotoImage(file=months_file)
month_img_ = w.create_image(51 * screen_x / 100, 71 * screen_y / 100, image=month_img)
# -- Create Buttons for main Page
# ---- Buttons Functions


def quit_fun():
    root.destroy()
    #sys.exit()


def transfer_page_start():
    w.delete(button1_window)
    w.delete(button2_window)
    w.delete(button3_window)
    w.delete(button4_window)
    # w.delete(button5_window)
    w.delete(TIME_img)
    w.delete(Logo)



    global transfer_page_detail, transfer_page_detail_
    transfer_page_detail = PhotoImage(file="images\Transfer_page.png")
    transfer_page_detail_ = w.create_image(screen_x / 4, screen_y / 8, anchor=NW, image=transfer_page_detail)
    global transfer_list
    transfer_list = Listbox(root, font=list_font)
    global scroll_t
    scroll_t = Scrollbar(root, orient=VERTICAL, command=transfer_list.yview)
    transfer_list['yscrollcommand'] = scroll_t.set
    global transfer_list_box
    transfer_list_box = w.create_window(4 * screen_x / 7, 5 * screen_y / 7, width=screen_x / 2, height=screen_y / 2,
                                      window=transfer_list)
    global scroll_box_t
    scroll_box_t = w.create_window(4 * screen_x / 7, 5 * screen_y / 7, width=screen_x / 2, height=screen_y / 2,
                                 window=scroll_t)
    global transfer_img
    transfer_img = PhotoImage(file="images\Transfer.png")
    button10 = Button(root, command=transfer_fun)
    button10.configure(image=transfer_img, width=3.6 * screen_x / 30, height=screen_y / 15, relief=FLAT)
    global button10_window
    button10_window = w.create_window(screen_x / 30, 6 * screen_y / 8, anchor=NW, window=button10)
    global Back3_img
    Back3_img = PhotoImage(file="images\Back.png")
    button11 = Button(root, command=transfer_page_stop)
    button11.configure(image=Back3_img, width=3.6 * screen_x / 30, height=screen_y / 15, relief=FLAT)
    global button11_window
    button11_window = w.create_window(screen_x / 30, 7 * screen_y / 8, anchor=NW, window=button11)
    global main_page_bool
    main_page_bool = False


def transfer_page_stop():
    w.delete(transfer_list_box)
    w.delete(scroll_box_t)
    w.delete(button10_window)
    w.delete(button11_window)
    w.delete(transfer_page_detail_)
    global button1_window, button2_window, button3_window, button4_window, Logo #button5_window, Logo
    button1_window = w.create_window(screen_x / 30, 7 * screen_y / 8, anchor=NW, window=button1)
    button2_window = w.create_window(screen_x / 30, 6 * screen_y / 8, anchor=NW, window=button2)
    button3_window = w.create_window(screen_x / 30, 5 * screen_y / 8, anchor=NW, window=button3)
    button4_window = w.create_window(screen_x / 30, 4 * screen_y / 8, anchor=NW, window=button4)
    #button5_window = w.create_window(screen_x / 30, 3 * screen_y / 8, anchor=NW, window=button5)
    Logo = w.create_image(screen_x / 2, screen_y / 6, anchor=CENTER, image=Logo_image)
    global main_page_bool
    main_page_bool = True


def transfer_fun():
    print("ccc")

def remove_page_start():
    w.delete(button1_window)
    w.delete(button2_window)
    w.delete(button3_window)
    w.delete(button4_window)
    #w.delete(button5_window)
    w.delete(TIME_img)
    w.delete(Logo)
    global remove_page_detail, remove_page_detail_
    remove_page_detail = PhotoImage(file="images\Remove_page.png")
    remove_page_detail_ = w.create_image(screen_x / 4, screen_y / 8, anchor=NW, image=remove_page_detail)
    global remove_list
    remove_list = Listbox(root, font=list_font)
    global scroll
    scroll = Scrollbar(root, orient=VERTICAL, command=remove_list.yview)
    remove_list['yscrollcommand'] = scroll.set
    global remove_list_box
    remove_list_box = w.create_window(4*screen_x/7, 5*screen_y/7, width=screen_x/2, height=screen_y/2, window=remove_list)
    global scroll_box
    scroll_box = w.create_window(4*screen_x/7, 5*screen_y/7, width=screen_x/2, height=screen_y/2, window=scroll)

    # Read Persons From File
    file_list = open("Persons\\Users_List.txt")
    total_user = len(file_list.read().split('\n'))
    file_list.close()
    file_list = open("Persons\\Users_List.txt")
    for i in range(0, total_user):
        person_ = str(file_list.readline())[11:]
        person.append(person_)
        remove_list.insert('end', person_)
    file_list.close()
    #########################

    global remove_img
    remove_img = PhotoImage(file="images\Remove.png")
    button8 = Button(root, command=remove_fun)
    button8.configure(image=remove_img, width=3.6 * screen_x / 30, height=screen_y / 15, relief=FLAT)
    global button8_window
    button8_window = w.create_window(screen_x / 30, 6 * screen_y / 8, anchor=NW, window=button8)
    global Back2_img
    Back2_img = PhotoImage(file="images\Back.png")
    button9 = Button(root, command=remove_page_stop)
    button9.configure(image=Back2_img, width=3.6 * screen_x / 30, height=screen_y / 15, relief=FLAT)
    global button9_window
    button9_window = w.create_window(screen_x / 30, 7 * screen_y / 8, anchor=NW, window=button9)
    global main_page_bool
    main_page_bool = False


def remove_page_stop():
    w.delete(remove_list_box)
    w.delete(scroll_box)
    w.delete(button8_window)
    w.delete(button9_window)
    w.delete(remove_page_detail_)
    global button1_window, button2_window, button3_window, button4_window, Logo #button5_window, Logo
    button1_window = w.create_window(screen_x / 30, 7 * screen_y / 8, anchor=NW, window=button1)
    button2_window = w.create_window(screen_x / 30, 6 * screen_y / 8, anchor=NW, window=button2)
    button3_window = w.create_window(screen_x / 30, 5 * screen_y / 8, anchor=NW, window=button3)
    button4_window = w.create_window(screen_x / 30, 4 * screen_y / 8, anchor=NW, window=button4)
    #button5_window = w.create_window(screen_x / 30, 3 * screen_y / 8, anchor=NW, window=button5)
    Logo = w.create_image(screen_x / 2, screen_y / 6, anchor=CENTER, image=Logo_image)
    global main_page_bool
    main_page_bool = True


def remove_fun():
    global person
    global remove_list
    person_ = person[int(remove_list.curselection()[0])]
    file_list = open("Persons\\Users_List.txt")
    list_1 = file_list.read()
    file_list.close()
    list_1 = list_1.replace(list_1[list_1.find(person_) - 11:list_1.find(person_) + len(person_)], "")
    file_list = open("Persons\\Users_List.txt", "w")
    file_list.write(list_1)
    file_list.close()
    for i in range(0, len(person)):
        person.pop(0)
    w.delete(remove_list_box)
    w.delete(scroll_box)
    w.delete(button8_window)
    w.delete(button9_window)
    remove_page_start()


def add_page_start():
    w.delete(DAY_img)
    w.delete(YEAR_img)
    w.delete(month_img_)
    w.delete(button1_window)
    w.delete(button2_window)
    w.delete(button3_window)
    w.delete(button4_window)
    # w.delete(button5_window)
    w.delete(TIME_img)
    w.delete(Logo)
    global save_page_detail, save_page_detail_
    save_page_detail = PhotoImage(file="images\Save_page.png")
    save_page_detail_ = w.create_image(screen_x / 4, screen_y / 8, anchor=NW, image=save_page_detail)
    global user
    user = Entry(w, font=name_font)
    global user_name_box
    user_name_box = w.create_window(5*screen_x / 9, screen_y /2, width=screen_x / 3, height=screen_y / 12, window=user)
    global rfid_logo, rfid_logo_
    rfid_logo = PhotoImage(file="images\RFID_Logo.png")
    rfid_logo_ = w.create_image(4*screen_x / 9, 5*screen_y /8, anchor=NW, image=rfid_logo)
    global Back1_img
    Back1_img = PhotoImage(file="images\Back.png")
    button7 = Button(root, command=add_page_stop)
    button7.configure(image=Back1_img, width=3.6 * screen_x / 30, height=screen_y / 15, relief=FLAT)
    global button7_window
    button7_window = w.create_window(screen_x / 30, 7 * screen_y / 8, anchor=NW, window=button7)
    global main_page_bool
    global add_page_bool
    main_page_bool = False
    add_page_bool = True


def add_page_stop():
    w.delete(button7_window)
    w.delete(user_name_box)
    w.delete(rfid_logo_)
    w.delete(save_page_detail_)
    global button1_window, button2_window, button3_window, button4_window, Logo # button5_window, Logo
    button1_window = w.create_window(screen_x / 30, 7 * screen_y / 8, anchor=NW, window=button1)
    button2_window = w.create_window(screen_x / 30, 6 * screen_y / 8, anchor=NW, window=button2)
    button3_window = w.create_window(screen_x / 30, 5 * screen_y / 8, anchor=NW, window=button3)
    button4_window = w.create_window(screen_x / 30, 4 * screen_y / 8, anchor=NW, window=button4)
    # button5_window = w.create_window(screen_x / 30, 3 * screen_y / 8, anchor=NW, window=button5)
    Logo = w.create_image(screen_x / 2, screen_y / 6, anchor=CENTER, image=Logo_image)
    global DAY_img, YEAR_img
    global month_img, month_img_
    year, day, time = persian_time()
    DAY_img = w.create_text(12 * screen_x / 20, 7 * screen_y / 10, text=day, font=date_font, fill='white')
    YEAR_img = w.create_text(8 * screen_x / 20, 7 * screen_y / 10, text=year, font=date_font, fill='white')
    months_file = "images\Months\\" + str(jdatetime.date.today())[5:7] + ".png"
    month_img = PhotoImage(file=months_file)
    month_img_ = w.create_image(51 * screen_x / 100, 71 * screen_y / 100, image=month_img)
    global main_page_bool
    global add_page_bool
    main_page_bool = True
    add_page_bool = False

def add_person():
    file_list = open("Persons\\Users_List.txt")
    len_of_file = len(file_list.read())
    file_list.close()
    file_list = open("Persons\\Users_List.txt", "a")
    global permanent_read_line
    if len_of_file == 0:
        file_list.write(permanent_read_line[2:12] + ":" + user.get())
    else:
        file_list.write("\n" + permanent_read_line[2:12] + ":" + user.get())
    file_list.close()
    global row
    wb = load_workbook(path)
    ws = wb.active
    ws["A" + str(row)] = user.get()
    ws["D" + str(row)] = permanent_read_line[2:12]
    wb.save(path)
    row = row + 1
    file_list = open("variables.txt", "w")
    file_list.write("ROW=" + str(row))
    file_list.close()
    permanent_read_line = ""
    w.delete(button6_window)
    user.delete(0, END)

# def vacation_fun():
#    root.destroy()


# ---- Create Buttons
# ------ Exit Button
Exit_img = PhotoImage(file="images\Exit.png")
button1 = Button(root, command=quit_fun)
button1.configure(image=Exit_img, width=3.6*screen_x/30, height=screen_y/15, relief=FLAT)
button1_window = w.create_window(screen_x/30, 7*screen_y/8, anchor=NW, window=button1)

# ------ Transfer data Button
Transfer_img = PhotoImage(file="images\Transfer.png")
button2 = Button(root, command=transfer_page_start)
button2.configure(image=Transfer_img, width=3.6*screen_x/30, height=screen_y/15, relief=FLAT)
button2_window = w.create_window(screen_x/30, 6*screen_y/8, anchor=NW, window=button2)

# ------ Remove data Button
Remove_img = PhotoImage(file="images\Remove_main.png")
button3 = Button(root, command=remove_page_start)
button3.configure(image=Remove_img, width=3.6*screen_x/30, height=screen_y/15, relief=FLAT)
button3_window = w.create_window(screen_x/30, 5*screen_y/8, anchor=NW, window=button3)

# ------ Add data Button
add_img = PhotoImage(file="images\Save_main.png")
button4 = Button(root, command=add_page_start)
button4.configure(image=add_img, width=3.6*screen_x/30, height=screen_y/15, relief=FLAT)
button4_window = w.create_window(screen_x/30, 4*screen_y/8, anchor=NW, window=button4)

# ------ Vacation data Button
# vacation_img = PhotoImage(file="images\Vacation.png")
# button5 = Button(root, command=vacation_fun)
# button5.configure(image=vacation_img, width=3.6*screen_x/30, height=screen_y/15, relief=FLAT)
# button5_window = w.create_window(screen_x/30, 3*screen_y/8, anchor=NW, window=button5)
# ALL Functions
# -- Change English Numbers to Persian


numbers = {'0': '۰', '1': '۱', '2': '۲', '3': '۳', '4': '۴', '5': '۵', '6': '۶', '7': '۷', '8': '۸', '9': '۹'}


def EN_2_PR(num):
    num_1 = list(num)
    for i in range(0, len(num)):
        num_1[i] = str(numbers[num[i]])
    ss = "".join(num_1)
    return ss

# -- Get Current Time (and Return Digital Clock Style)


def persian_time():
    year = EN_2_PR(jdatetime.datetime.now().strftime("%Y"))
    day = EN_2_PR(jdatetime.datetime.now().strftime("%d"))
    hour = EN_2_PR(jdatetime.datetime.now().strftime("%H"))
    minute = EN_2_PR(jdatetime.datetime.now().strftime("%M"))
    second = EN_2_PR(jdatetime.datetime.now().strftime("%S"))
    time = hour + ":" + minute + ":" + second
    return year, day, time

# -- Update main page Date information


def update_date():
    global DAY_img, YEAR_img
    global month_img, month_img_
    w.delete(DAY_img)
    w.delete(YEAR_img)
    w.delete(month_img_)
    year, day, time = persian_time()
    DAY_img = w.create_text(12 * screen_x / 20, 7 * screen_y / 10, text=day, font=date_font, fill='white')
    YEAR_img = w.create_text(8 * screen_x / 20, 7 * screen_y / 10, text=year, font=date_font, fill='white')
    months_file = "images\Months\\" + str(jdatetime.date.today())[5:7] + ".png"
    month_img = PhotoImage(file=months_file)
    month_img_ = w.create_image(51 * screen_x / 100, 71 * screen_y / 100, image=month_img)
    root.after(60000, update_date)
# -- Get

def create_file():
    global today
    tomorrow = today
    today = jdatetime.date.today()
    global path
    path = "data\\" + str(today) + ".xlsx"
    if tomorrow != today or not(exists(path)):
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "Person"; ws['B1'] = "Enter"; ws['C1'] = "Exit"; ws['D1'] = "Card Num"
        ws['A1'].font = ws['A1'].font.copy(bold=True, italic=True)
        ws['B1'].font = ws['B1'].font.copy(bold=True, italic=True)
        ws['C1'].font = ws['C1'].font.copy(bold=True, italic=True)
        ws['D1'].font = ws['C1'].font.copy(bold=True, italic=True)

        file_list = open("Persons\\Users_List.txt")
        total_user = len(file_list.read().split('\n'))
        file_list.close()
        file_list = open("Persons\\Users_List.txt")
        global row
        for i in range(0, total_user):
            person_ = str(file_list.readline())
            pos_A = "A" + str(i+2)
            pos_D = "D" + str(i+2)
            ws[pos_A] = person_[11:]
            ws[pos_D] = person_[0:10]
            row = i+2
        file_list.close()
        file_list = open("variables.txt", "w")
        file_list.write("ROW=" + str(row))
        file_list.close()
        wb.save(path)
    else:
        None
    root.after(60000, create_file)


def search_person(card_num):
    pos = 0
    enter_ = ""
    exit_ = ""
    person_name = ""
    wb = load_workbook(path)
    ws = wb.active
    file_list = open("Persons\\Users_List.txt")
    total_user = len(file_list.read().split('\n'))
    file_list.close()
    for i in range(0, total_user):
        pos_D = "D" + str(i + 2)
        cardnum_buffer = ws[pos_D].value
        if cardnum_buffer == card_num:
            pos = i+2
            person_name = ws["A" + str(i+2)].value
            enter_ = ws["B" + str(i+2)].value
            exit_ = ws["C" + str(i+2)].value

    return pos, enter_, exit_, person_name


def rollcall(card_num):
    global info_text, info_text_
    global info_img, info_img_
    info_img = PhotoImage(file="images\info.png")
    info_img_ = w.create_image(screen_x / 2, screen_y / 2, image=info_img)
    file_list = open("Persons\\Users_List.txt")
    total_user = len(file_list.read().split('\n'))
    file_list.close()
    file_list = open("Persons\\Users_List.txt")
    file_ = file_list.read()
    if file_.find(card_num) != -1:
        pos, enter_, exit_, person_name = search_person(card_num)
        year, day, time = persian_time()
        wb = load_workbook(path)
        ws = wb.active
        if enter_ == None and exit_ == None:
            ws["B"+str(pos)] = time
            enter_ = time
        elif enter_ != None and exit_ == None:
            ws["C" + str(pos)] = time
            exit_ = time
        wb.save(path)
        if exit_ == None:
            exit_ = "----"
        info_text = person_name + "\n\n" + card_num + "\n\n\n" + enter_ + "\n\n" + exit_
        info_text_ = w.create_text(4*screen_x / 14, 2*screen_y / 7, anchor=NW, text=info_text,
                                   font=info_font, fill='Black')
    else:
        None
        info_text = "Unknown CARD"
        info_text_ = w.create_text(4 * screen_x / 14, 2 * screen_y / 7, anchor=NW, text=info_text, font=info_font,
                                   fill='Black')
# -- Show main page on GUI


def remove_info():
    w.delete(info_text_)
    w.delete(info_img_)
    global main_page_bool
    main_page_bool = True


def remove_card_error():
    w.delete(card_error_img)


def main_loop():
    global main_page_bool
    if main_page_bool:
        global TIME_img
        global DAY_img
        global YEAR_img
        w.delete(TIME_img)
        year, day, time = persian_time()
        TIME_img = w.create_text(screen_x / 2, screen_y / 2, text=time, font=time_font, fill='white')
        ## Wait For RF Card
        read_line = str(rf.readline())
        if len(read_line) > 5:
            rollcall(read_line[2:12])
            main_page_bool = False
            root.after(2000, remove_info)
    global add_page_bool
    if add_page_bool:
        global permanent_read_line
        read_line = str(rf.readline())
        global user
        if len(read_line) > 5:
            permanent_read_line = read_line
            if len(user.get()) > 0:
                file_list = open("Persons\\Users_List.txt")
                file = file_list.read()
                file_list.close()
                if file.find(permanent_read_line[2:12]) == -1:
                    global save_img
                    save_img = PhotoImage(file="images\Save.png")
                    button6 = Button(root, command=add_person)
                    button6.configure(image=save_img, width=3.6 * screen_x / 30, height=screen_y / 15, relief=FLAT)
                    global button6_window
                    button6_window = w.create_window(screen_x / 30, 6 * screen_y / 8, anchor=NW, window=button6)
                elif file.find(permanent_read_line[2:12]) != -1:
                    global card_error
                    global card_error_img
                    card_error = PhotoImage(file="images\Card_error.png")
                    card_error_img = w.create_image(5*screen_x / 9, 15*screen_y /16, image=card_error)
                    root.after(3000, remove_card_error)
    root.after(100, main_loop)

create_file()
root.after(60000, create_file)
root.after(60000, update_date)
root.after(10, main_loop)
root.mainloop()
